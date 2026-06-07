import httpx, random, sys, io
from openpyxl import load_workbook
BASE = "https://construction-accountant.onrender.com/api"
ok = True
def ck(n, c, x=""):
    global ok; ok &= c; print(("PASS" if c else "FAIL"), n, x)
with httpx.Client(timeout=60) as c:
    email = f"livefeat_{random.randint(100000,999999)}@x.com"
    tok = c.post(BASE+"/auth/register", json={"email": email, "password": "secret123", "name": "LF"}).json()["token"]
    H = {"Authorization": "Bearer " + tok}
    csv = "Огноо/Date,Төрөл/Type,Дүн/Amount,Ангилал,Тайлбар\n2026-06-05,Орлого/income,5000000,Борлуулалт,Гэрээ\n2026-06-06,Зарлага/expense,555000,Материал,Цемент\n".encode("utf-8")
    r = c.post(BASE+"/import", headers=H, files={"file": ("led.csv", csv, "text/csv")}, data={"kind":"ledger","month":"2026-06","confirm":"false"})
    ck("import_preview", r.status_code==200 and r.json()["total"]==2, str(r.status_code))
    r = c.post(BASE+"/import", headers=H, files={"file": ("led.csv", csv, "text/csv")}, data={"kind":"ledger","month":"2026-06","confirm":"true"})
    ck("import_confirm", r.status_code==200 and r.json()["imported"]==2)
    r = c.get(BASE+"/report/monthly?month=6&year=2026", headers=H)
    ck("monthly_report", r.status_code==200 and "spreadsheet" in r.headers.get("content-type",""), f"{len(r.content)}b")
    sheets = load_workbook(io.BytesIO(r.content)).sheetnames
    ck("report_5_sheets", len(sheets)==5)
    r = c.get(BASE+"/import/template?kind=materials", headers=H)
    ck("template", r.status_code==200)
print("\n=>", "ALL PASS" if ok else "FAILED")
sys.exit(0 if ok else 1)
