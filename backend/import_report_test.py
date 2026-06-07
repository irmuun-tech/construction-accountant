import httpx, random, sys, io
from openpyxl import Workbook, load_workbook
BASE = "http://127.0.0.1:8001/api"
ok = True
def check(n, c, extra=""):
    global ok; ok &= c; print(("PASS" if c else "FAIL"), n, extra)

with httpx.Client(timeout=30) as c:
    email = f"imp_{random.randint(100000,999999)}@x.com"
    tok = c.post(BASE+"/auth/register", json={"email": email, "password": "secret123", "name": "Imp"}).json()["token"]
    H = {"Authorization": "Bearer " + tok}

    # --- Ledger import via CSV: preview then confirm ---
    csv_led = "Огноо/Date,Төрөл/Type,Дүн/Amount,Ангилал/Category,Тайлбар/Description\n2026-06-05,Орлого/income,5000000,Борлуулалт,Гэрээ\n2026-06-06,Зарлага/expense,555000,Материал,Цемент\n".encode("utf-8")
    r = c.post(BASE+"/import", headers=H, files={"file": ("led.csv", csv_led, "text/csv")}, data={"kind":"ledger","month":"2026-06","confirm":"false"})
    j = r.json(); check("ledger_preview", r.status_code==200 and j["total"]==2 and len(j["preview"])==2, str(j.get("total")))
    check("ledger_type_detect", j["preview"][0]["type"]=="income" and j["preview"][1]["type"]=="outcome")
    r = c.post(BASE+"/import", headers=H, files={"file": ("led.csv", csv_led, "text/csv")}, data={"kind":"ledger","month":"2026-06","confirm":"true"})
    check("ledger_import", r.status_code==200 and r.json()["imported"]==2)

    # --- Materials import via real XLSX (with quantity -> stock) ---
    wb = Workbook(); ws = wb.active
    ws.append(["Нэр/Name","Ангилал/Category","Нэгж/Unit","Нэгж үнэ/Price","Нийлүүлэгч/Supplier","Тоо хэмжээ/Quantity"])
    ws.append(["Цемент 500","Цемент","уут",18500,"Хөтөл",40])
    ws.append(["Арматур 12мм","Төмөр","м",3200,"Дархан",455])
    bio = io.BytesIO(); wb.save(bio); xlsx_bytes = bio.getvalue()
    r = c.post(BASE+"/import", headers=H, files={"file": ("mat.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"kind":"materials","month":"2026-06","confirm":"false"})
    j = r.json(); check("materials_preview_xlsx", r.status_code==200 and j["total"]==2, str(j.get("total")))
    check("materials_price_parsed", j["preview"][0]["price"]==18500 and j["preview"][0]["quantity"]==40)
    r = c.post(BASE+"/import", headers=H, files={"file": ("mat.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"kind":"materials","month":"2026-06","confirm":"true"})
    j = r.json(); check("materials_import", r.status_code==200 and j["imported"]==2 and j["materials_added"]==2 and j["stock_added"]==2, str(j))

    # --- Verify data landed ---
    mats = c.get(BASE+"/materials", headers=H).json(); check("materials_in_db", len(mats)==2)
    txns = c.get(BASE+"/transactions", headers=H).json(); check("txns_in_db", len(txns)==2)
    stock = c.get(BASE+"/stock?month=6&year=2026", headers=H).json(); check("stock_in_db", len(stock)==2)

    # --- Templates ---
    for k in ("ledger","materials"):
        r = c.get(BASE+f"/import/template?kind={k}", headers=H)
        check(f"template_{k}", r.status_code==200 and "spreadsheet" in r.headers.get("content-type",""), f"{len(r.content)}b")

    # --- Monthly report (multi-sheet) ---
    r = c.get(BASE+"/report/monthly?month=6&year=2026", headers=H)
    check("report_status", r.status_code==200 and "spreadsheet" in r.headers.get("content-type",""), f"{len(r.content)}b")
    rb = load_workbook(io.BytesIO(r.content))
    check("report_sheets", len(rb.sheetnames)==5, str(rb.sheetnames))
    # summary income should be 5,000,000
    summ = rb["Хураангуй-Summary"]
    vals = [summ.cell(row=rr, column=2).value for rr in range(1,12)]
    check("report_income_total", 5000000 in vals, str(vals))

    # bad file -> helpful 400
    r = c.post(BASE+"/import", headers=H, files={"file": ("x.csv", b"foo,bar\n1,2\n", "text/csv")}, data={"kind":"ledger","confirm":"false"})
    check("ledger_missing_amount_400", r.status_code==400)

passed = "ALL PASS" if ok else "SOME FAILED"
print("\n=>", passed)
sys.exit(0 if ok else 1)
