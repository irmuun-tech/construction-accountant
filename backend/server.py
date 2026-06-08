"""
Барилгын Нягтлан · Construction Accountant — backend API
FastAPI + MongoDB. Self-contained email/password auth (bcrypt + JWT).

Sections: auth, materials (search), stock (monthly stocktaking + Excel),
transactions (income/outcome), loans (reminders), dashboard.
"""
from fastapi import FastAPI, APIRouter, Header, HTTPException, UploadFile, File, Form, Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO, StringIO
import csv
import json
import re
import time
from collections import defaultdict
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ---------- Config ----------
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'construction_accountant')
JWT_SECRET = os.environ.get('JWT_SECRET', 'dev-only-change-me')
JWT_ALGO = 'HS256'
JWT_DAYS = 30
CORS_ORIGINS = os.environ.get('CORS_ORIGINS', '*')
# Signup gate. Set INVITE_CODE on the host (one or more, comma-separated) to require a code.
INVITE_CODES = [c.strip() for c in os.environ.get('INVITE_CODE', '').split(',') if c.strip()]
# Registration is CLOSED by default for safety (no stranger can self-register).
# Open it by setting REGISTRATION_OPEN=true, or by setting INVITE_CODE (code required).
REGISTRATION_OPEN = os.environ.get('REGISTRATION_OPEN', 'false').strip().lower() in ('1', 'true', 'yes', 'on')

# --- Simple in-memory rate limiting (per process) ---
_RL = defaultdict(list)
def _client_ip(request: Request) -> str:
    fwd = request.headers.get('x-forwarded-for')
    return (fwd.split(',')[0].strip() if fwd else (request.client.host if request.client else 'unknown'))
def _rate_limit(key: str, max_calls: int, window: int = 60):
    now = time.time()
    hits = [t for t in _RL[key] if now - t < window]
    if len(hits) >= max_calls:
        raise HTTPException(status_code=429, detail="Хэт олон оролдлого, түр хүлээнэ үү / Too many requests, please slow down")
    hits.append(now)
    _RL[key] = hits

client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

app = FastAPI(title="Construction Accountant API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str
    invite_code: Optional[str] = ""

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    user_id: str
    email: str
    name: str
    company: Optional[str] = ""

class AuthResponse(BaseModel):
    token: str
    user: UserResponse

class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    company: Optional[str] = None

class PasswordChange(BaseModel):
    current_password: str
    new_password: str = Field(min_length=6)

class MaterialCreate(BaseModel):
    name: str
    category: str = ""
    supplier: str = ""
    price: float = 0
    unit: str = "ш"
    min_stock: float = 0
    description: Optional[str] = ""
    image_url: Optional[str] = ""

class MaterialResponse(BaseModel):
    material_id: str
    name: str
    category: str
    supplier: str
    price: float
    unit: str
    min_stock: float = 0
    description: str
    image_url: Optional[str] = ""
    created_at: datetime

class StockEntryCreate(BaseModel):
    material_id: str
    quantity: float
    date: str  # YYYY-MM-DD

class StockEntryResponse(BaseModel):
    entry_id: str
    material_id: str
    material_name: str
    unit: str = ""
    quantity: float
    date: str
    month: int
    year: int
    created_at: datetime

class TransactionCreate(BaseModel):
    type: str            # "income" or "outcome"
    category: str = ""
    amount: float
    description: Optional[str] = ""
    date: str            # YYYY-MM-DD
    # Optional material-purchase details (нэгжийн үнэ × ширхэг)
    material_id: Optional[str] = ""
    quantity: Optional[float] = None
    unit_price: Optional[float] = None

class TransactionResponse(BaseModel):
    transaction_id: str
    type: str
    category: str
    amount: float
    description: str
    date: str
    created_at: datetime
    material_id: Optional[str] = ""
    material_name: Optional[str] = ""
    quantity: Optional[float] = None
    unit_price: Optional[float] = None

class LoanCreate(BaseModel):
    name: str
    amount: float
    interest_rate: float = 0
    payment_amount: float = 0
    frequency: str = "monthly"   # "monthly" or "weekly"
    start_date: str              # YYYY-MM-DD

class LoanResponse(BaseModel):
    loan_id: str
    name: str
    amount: float
    interest_rate: float
    payment_amount: float
    frequency: str
    start_date: str
    next_payment_date: str
    status: str
    created_at: datetime

class DashboardStats(BaseModel):
    total_materials: int
    total_stock_entries: int
    monthly_income: float
    monthly_outcome: float
    active_loans: int
    upcoming_payments: int

# ==================== AUTH HELPERS ====================

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False

def make_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=JWT_DAYS)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")
    token = authorization.replace("Bearer ", "", 1)
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        user_id = payload.get("sub")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Session expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=AuthResponse)
async def register(body: RegisterRequest, request: Request):
    _rate_limit("reg:" + _client_ip(request), 15, 60)
    if INVITE_CODES:
        if (body.invite_code or "").strip() not in INVITE_CODES:
            raise HTTPException(status_code=403, detail="Урилгын код буруу эсвэл дутуу байна / Invalid or missing invite code")
    elif not REGISTRATION_OPEN:
        raise HTTPException(status_code=403, detail="Шинэ бүртгэл хаалттай байна. Админтай холбогдоно уу / Registration is closed — contact the admin")
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=409, detail="Энэ имэйл бүртгэлтэй байна / Email already registered")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    await db.users.insert_one({
        "user_id": user_id,
        "email": body.email.lower(),
        "name": body.name,
        "company": "",
        "password": hash_password(body.password),
        "created_at": datetime.now(timezone.utc)
    })
    token = make_token(user_id)
    return AuthResponse(token=token, user=UserResponse(user_id=user_id, email=body.email.lower(), name=body.name, company=""))

@api_router.post("/auth/login", response_model=AuthResponse)
async def login(body: LoginRequest, request: Request):
    _rate_limit("login:" + _client_ip(request), 30, 60)
    user = await db.users.find_one({"email": body.email.lower()})
    if not user or not verify_password(body.password, user.get("password", "")):
        raise HTTPException(status_code=401, detail="Имэйл эсвэл нууц үг буруу / Wrong email or password")
    token = make_token(user["user_id"])
    return AuthResponse(token=token, user=UserResponse(
        user_id=user["user_id"], email=user["email"], name=user["name"], company=user.get("company", "")))

@api_router.get("/auth/me", response_model=UserResponse)
async def me(authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    return UserResponse(**user)

@api_router.patch("/auth/profile", response_model=UserResponse)
async def update_profile(body: ProfileUpdate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": updates})
    user = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password": 0})
    return UserResponse(**user)

@api_router.post("/auth/change-password")
async def change_password(body: PasswordChange, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    full = await db.users.find_one({"user_id": user["user_id"]})
    if not full or not verify_password(body.current_password, full.get("password", "")):
        raise HTTPException(status_code=400, detail="Одоогийн нууц үг буруу / Current password is incorrect")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"password": hash_password(body.new_password)}})
    return {"message": "Нууц үг шинэчлэгдлээ / Password updated"}

# ==================== MATERIALS ====================

@api_router.post("/materials", response_model=MaterialResponse)
async def create_material(material: MaterialCreate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    doc = {
        "material_id": f"mat_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "name": material.name, "category": material.category, "supplier": material.supplier,
        "price": material.price, "unit": material.unit, "min_stock": material.min_stock,
        "description": material.description, "image_url": material.image_url or "",
        "created_at": datetime.now(timezone.utc)
    }
    await db.materials.insert_one(doc)
    return MaterialResponse(**doc)

@api_router.get("/materials", response_model=List[MaterialResponse])
async def get_materials(
    search: Optional[str] = None, category: Optional[str] = None, supplier: Optional[str] = None,
    min_price: Optional[float] = None, max_price: Optional[float] = None,
    authorization: Optional[str] = Header(None)
):
    user = await get_current_user(authorization)
    query = {"user_id": user["user_id"]}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    if category:
        query["category"] = category
    if supplier:
        query["supplier"] = {"$regex": supplier, "$options": "i"}
    if min_price is not None or max_price is not None:
        query["price"] = {}
        if min_price is not None:
            query["price"]["$gte"] = min_price
        if max_price is not None:
            query["price"]["$lte"] = max_price
    materials = await db.materials.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [MaterialResponse(**m) for m in materials]

@api_router.get("/materials/categories")
async def get_categories(authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    cats = await db.materials.distinct("category", {"user_id": user["user_id"]})
    return {"categories": [c for c in cats if c]}

@api_router.put("/materials/{material_id}", response_model=MaterialResponse)
async def update_material(material_id: str, material: MaterialCreate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.materials.update_one(
        {"material_id": material_id, "user_id": user["user_id"]},
        {"$set": material.model_dump()}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    doc = await db.materials.find_one({"material_id": material_id}, {"_id": 0})
    return MaterialResponse(**doc)

@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.materials.delete_one({"material_id": material_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    return {"message": "deleted"}

# ==================== STOCK (stocktaking) ====================

@api_router.post("/stock", response_model=StockEntryResponse)
async def create_stock_entry(entry: StockEntryCreate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    material = await db.materials.find_one({"material_id": entry.material_id, "user_id": user["user_id"]}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    try:
        date_obj = datetime.strptime(entry.date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    doc = {
        "entry_id": f"stock_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "material_id": entry.material_id,
        "material_name": material["name"],
        "unit": material.get("unit", ""),
        "quantity": entry.quantity,
        "date": entry.date, "month": date_obj.month, "year": date_obj.year,
        "created_at": datetime.now(timezone.utc)
    }
    await db.stock_entries.insert_one(doc)
    return StockEntryResponse(**doc)

@api_router.get("/stock", response_model=List[StockEntryResponse])
async def get_stock(month: Optional[int] = None, year: Optional[int] = None, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    query = {"user_id": user["user_id"]}
    if month:
        query["month"] = month
    if year:
        query["year"] = year
    entries = await db.stock_entries.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    return [StockEntryResponse(**e) for e in entries]

@api_router.get("/stock/export")
async def export_stock(month: Optional[int] = None, year: Optional[int] = None, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    query = {"user_id": user["user_id"]}
    if month:
        query["month"] = month
    if year:
        query["year"] = year
    entries = await db.stock_entries.find(query, {"_id": 0}).sort("date", 1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Тооллого Stock"
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Огноо/Date", "Материал/Material", "Нэгж/Unit", "Тоо/Qty", "Сар/Month", "Жил/Year"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center")
    for i, e in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=e["date"])
        ws.cell(row=i, column=2, value=e["material_name"])
        ws.cell(row=i, column=3, value=e.get("unit", ""))
        ws.cell(row=i, column=4, value=e["quantity"])
        ws.cell(row=i, column=5, value=e["month"])
        ws.cell(row=i, column=6, value=e["year"])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = width + 3

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"stocktaking_{year or 'all'}_{month or 'all'}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@api_router.delete("/stock/{entry_id}")
async def delete_stock(entry_id: str, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.stock_entries.delete_one({"entry_id": entry_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock entry not found")
    return {"message": "deleted"}

# ==================== TRANSACTIONS (income / outcome) ====================

@api_router.post("/transactions", response_model=TransactionResponse)
async def create_transaction(t: TransactionCreate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    if t.type not in ("income", "outcome"):
        raise HTTPException(status_code=400, detail="type must be income or outcome")
    material_name = ""
    if t.material_id:
        mat = await db.materials.find_one({"material_id": t.material_id, "user_id": user["user_id"]}, {"_id": 0})
        if mat:
            material_name = mat["name"]
    doc = {
        "transaction_id": f"txn_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "type": t.type, "category": t.category, "amount": t.amount,
        "description": t.description, "date": t.date, "created_at": datetime.now(timezone.utc),
        "material_id": t.material_id or "", "material_name": material_name,
        "quantity": t.quantity, "unit_price": t.unit_price
    }
    await db.transactions.insert_one(doc)
    return TransactionResponse(**doc)

@api_router.get("/transactions", response_model=List[TransactionResponse])
async def get_transactions(
    type: Optional[str] = None, category: Optional[str] = None,
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    user = await get_current_user(authorization)
    query = {"user_id": user["user_id"]}
    if type:
        query["type"] = type
    if category:
        query["category"] = category
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = start_date
        if end_date:
            query["date"]["$lte"] = end_date
    txns = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    return [TransactionResponse(**t) for t in txns]

@api_router.get("/transactions/summary")
async def transaction_summary(month: Optional[int] = None, year: Optional[int] = None, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    date_filter = {}
    if month and year:
        start = f"{year}-{month:02d}-01"
        end = f"{year+1}-01-01" if month == 12 else f"{year}-{(month+1):02d}-01"
        date_filter = {"$gte": start, "$lt": end}
    elif year:
        date_filter = {"$gte": f"{year}-01-01", "$lt": f"{year+1}-01-01"}
    query = {"user_id": user["user_id"]}
    if date_filter:
        query["date"] = date_filter
    txns = await db.transactions.find(query, {"_id": 0}).to_list(20000)
    income_by, outcome_by, total_in, total_out = {}, {}, 0.0, 0.0
    for t in txns:
        cat = t.get("category") or "—"
        if t["type"] == "income":
            total_in += t["amount"]; income_by[cat] = income_by.get(cat, 0) + t["amount"]
        else:
            total_out += t["amount"]; outcome_by[cat] = outcome_by.get(cat, 0) + t["amount"]
    return {
        "total_income": total_in, "total_outcome": total_out, "net": total_in - total_out,
        "income_by_category": income_by, "outcome_by_category": outcome_by
    }

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.transactions.delete_one({"transaction_id": transaction_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "deleted"}

# ==================== LOANS ====================

def _next_payment(start_date: str, frequency: str) -> str:
    start = datetime.strptime(start_date, "%Y-%m-%d")
    delta = timedelta(days=7) if frequency == "weekly" else timedelta(days=30)
    nxt = start
    today = datetime.now()
    # advance to the first due date that is today or later
    while nxt < today:
        nxt = nxt + delta
    return nxt.strftime("%Y-%m-%d")

@api_router.post("/loans", response_model=LoanResponse)
async def create_loan(loan: LoanCreate, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    try:
        next_payment = _next_payment(loan.start_date, loan.frequency)
    except ValueError:
        raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
    doc = {
        "loan_id": f"loan_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "name": loan.name, "amount": loan.amount, "interest_rate": loan.interest_rate,
        "payment_amount": loan.payment_amount, "frequency": loan.frequency,
        "start_date": loan.start_date, "next_payment_date": next_payment,
        "status": "active", "created_at": datetime.now(timezone.utc)
    }
    await db.loans.insert_one(doc)
    return LoanResponse(**doc)

@api_router.get("/loans", response_model=List[LoanResponse])
async def get_loans(status: Optional[str] = None, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    query = {"user_id": user["user_id"]}
    if status:
        query["status"] = status
    loans = await db.loans.find(query, {"_id": 0}).sort("next_payment_date", 1).to_list(2000)
    return [LoanResponse(**l) for l in loans]

@api_router.get("/loans/upcoming", response_model=List[LoanResponse])
async def upcoming_loans(days: int = 7, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    today = datetime.now().strftime("%Y-%m-%d")
    future = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    query = {"user_id": user["user_id"], "status": "active", "next_payment_date": {"$gte": today, "$lte": future}}
    loans = await db.loans.find(query, {"_id": 0}).sort("next_payment_date", 1).to_list(500)
    return [LoanResponse(**l) for l in loans]

@api_router.post("/loans/{loan_id}/pay", response_model=LoanResponse)
async def pay_loan(loan_id: str, authorization: Optional[str] = Header(None)):
    """Record a payment: advance next_payment_date by one period and reduce the balance."""
    user = await get_current_user(authorization)
    loan = await db.loans.find_one({"loan_id": loan_id, "user_id": user["user_id"]}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    freq = loan.get("frequency", "monthly")
    delta = timedelta(days=7) if freq == "weekly" else timedelta(days=30)
    nxt = (datetime.strptime(loan["next_payment_date"], "%Y-%m-%d") + delta).strftime("%Y-%m-%d")
    # Amortize: the payment first covers this period's interest, the rest reduces the balance.
    # (rate == 0 → interest is 0 → behaves exactly like a flat payment, so old loans are unaffected.)
    balance = float(loan.get("amount", 0))
    rate = float(loan.get("interest_rate", 0) or 0)
    payment = float(loan.get("payment_amount", 0) or 0)
    periods_per_year = 52 if freq == "weekly" else 12
    interest = round(balance * (rate / 100.0) / periods_per_year, 2) if rate > 0 else 0.0
    principal = payment - interest
    new_amount = round(max(0.0, balance - principal), 2)
    status = "paid" if new_amount <= 0 else "active"
    await db.loans.update_one({"loan_id": loan_id, "user_id": user["user_id"]},
                              {"$set": {"next_payment_date": nxt, "amount": new_amount, "status": status}})
    loan.update({"next_payment_date": nxt, "amount": new_amount, "status": status})
    return LoanResponse(**loan)

@api_router.patch("/loans/{loan_id}/status")
async def update_loan_status(loan_id: str, status: str, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.loans.update_one({"loan_id": loan_id, "user_id": user["user_id"]}, {"$set": {"status": status}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Loan not found")
    return {"message": "updated"}

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    res = await db.loans.delete_one({"loan_id": loan_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Loan not found")
    return {"message": "deleted"}

# ==================== DASHBOARD ====================

@api_router.get("/dashboard", response_model=DashboardStats)
async def dashboard(authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    uid_q = {"user_id": user["user_id"]}
    now = datetime.now()
    start = f"{now.year}-{now.month:02d}-01"
    end = f"{now.year+1}-01-01" if now.month == 12 else f"{now.year}-{(now.month+1):02d}-01"

    total_materials = await db.materials.count_documents(uid_q)
    total_stock_entries = await db.stock_entries.count_documents(uid_q)
    txns = await db.transactions.find({**uid_q, "date": {"$gte": start, "$lt": end}}, {"_id": 0}).to_list(20000)
    monthly_income = sum(t["amount"] for t in txns if t["type"] == "income")
    monthly_outcome = sum(t["amount"] for t in txns if t["type"] == "outcome")
    active_loans = await db.loans.count_documents({**uid_q, "status": "active"})
    today = now.strftime("%Y-%m-%d")
    future = (now + timedelta(days=7)).strftime("%Y-%m-%d")
    upcoming = await db.loans.count_documents({**uid_q, "status": "active", "next_payment_date": {"$gte": today, "$lte": future}})

    return DashboardStats(
        total_materials=total_materials, total_stock_entries=total_stock_entries,
        monthly_income=monthly_income, monthly_outcome=monthly_outcome,
        active_loans=active_loans, upcoming_payments=upcoming
    )

@api_router.get("/backup")
async def backup_data(authorization: Optional[str] = Header(None)):
    """Download ALL of the user's data as a single JSON file (self-serve backup)."""
    user = await get_current_user(authorization)
    uid_q = {"user_id": user["user_id"]}
    materials = await db.materials.find(uid_q, {"_id": 0}).to_list(100000)
    stock = await db.stock_entries.find(uid_q, {"_id": 0}).to_list(100000)
    transactions = await db.transactions.find(uid_q, {"_id": 0}).to_list(100000)
    loans = await db.loans.find(uid_q, {"_id": 0}).to_list(100000)
    payload = {
        "app": "construction-accountant",
        "backup_version": 1,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "user": {"email": user.get("email"), "name": user.get("name"), "company": user.get("company", "")},
        "counts": {"materials": len(materials), "stock_entries": len(stock),
                   "transactions": len(transactions), "loans": len(loans)},
        "materials": materials, "stock_entries": stock,
        "transactions": transactions, "loans": loans,
    }
    data = json.dumps(payload, ensure_ascii=False, default=str, indent=2).encode("utf-8")
    fname = f"backup_{datetime.now().strftime('%Y-%m-%d')}.json"
    return StreamingResponse(BytesIO(data), media_type="application/json",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@api_router.get("/health")
async def health():
    try:
        await db.command("ping")
        return {"status": "ok", "db": "connected"}
    except Exception as e:
        return {"status": "degraded", "db": str(e)}

# ==================== FILE → EXCEL CONVERTER ====================

def _decode_bytes(b: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="replace")

def _coerce(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    s = str(v); t = s.strip()
    if t and len(t) <= 15 and t not in ("-", "."):
        try:
            if t.lstrip("-").isdigit():
                return int(t)
            return float(t)
        except Exception:
            pass
    return s

def _rows_from_file(content: bytes, ext: str):
    text = _decode_bytes(content)
    stripped = text.lstrip()
    if ext == "json" or stripped[:1] in "[{":
        try:
            data = json.loads(text)
        except Exception:
            data = None
        if data is not None:
            if isinstance(data, list):
                if data and all(isinstance(x, dict) for x in data):
                    headers = []
                    for x in data:
                        for k in x.keys():
                            if k not in headers:
                                headers.append(k)
                    return [headers] + [[x.get(h, "") for h in headers] for x in data]
                if data and all(isinstance(x, list) for x in data):
                    return [list(r) for r in data]
                return [["value"]] + [[x] for x in data]
            if isinstance(data, dict):
                return [["key", "value"]] + [[k, v] for k, v in data.items()]
    # delimited / plain text
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:
        if "\t" in sample:
            delim = "\t"
        elif ";" in sample:
            delim = ";"
    reader = csv.reader(StringIO(text), delimiter=delim)
    return [r for r in reader if any(str(c).strip() for c in r)]

@api_router.post("/convert")
async def convert_to_excel(file: UploadFile = File(...), authorization: Optional[str] = Header(None)):
    """Convert an uploaded CSV / TSV / JSON / text file into a styled .xlsx."""
    await get_current_user(authorization)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Хоосон файл / Empty file")
    if len(content) > 6_000_000:
        raise HTTPException(status_code=413, detail="Файл хэт том (6MB дээш) / File too large")
    fname = file.filename or "file"
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    if ext in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Аль хэдийн Excel файл байна / Already an Excel file")
    try:
        rows = _rows_from_file(content, ext)
    except Exception:
        rows = []
    if not rows:
        raise HTTPException(status_code=400, detail="Хүснэгт мэдээлэл олдсонгүй / No tabular data found")

    wb = Workbook(); ws = wb.active
    ws.title = (fname.rsplit(".", 1)[0] or "Sheet1")[:31]
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ncols = max((len(r) for r in rows), default=1)
    widths = [10] * ncols
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=_coerce(val))
            if i == 1:
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ln = len(str(cell.value)) if cell.value is not None else 0
            if j - 1 < ncols:
                widths[j - 1] = max(widths[j - 1], ln)
    for idx in range(ncols):
        ws.column_dimensions[ws.cell(row=1, column=idx + 1).column_letter].width = min(widths[idx] + 3, 60)
    if len(rows) > 1:
        ws.freeze_panes = "A2"

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    base = fname.rsplit(".", 1)[0] or "converted"
    safe = "".join(c if (ord(c) < 128 and c not in '\\/:*?"<>|') else "_" for c in base).strip() or "converted"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}.xlsx"', "X-Rows": str(len(rows))}
    )

# ==================== IMPORT (file -> records) + MONTHLY REPORT ====================

HEADER_KEYS = {
    'date': ['date', 'огноо'],
    'type': ['type', 'төрөл'],
    'amount': ['amount', 'дүн', 'үнэ', 'мөнгө', 'төгрөг', 'нийт', 'total', 'sum'],
    'category': ['category', 'ангилал'],
    'description': ['description', 'тайлбар', 'утга', 'гүйлгээ', 'note', 'тэмдэглэл'],
    'name': ['name', 'нэр', 'материал', 'material'],
    'unit': ['unit', 'нэгж'],
    'price': ['price', 'үнэ', 'unit price'],
    'supplier': ['supplier', 'нийлүүлэгч'],
    'quantity': ['quantity', 'тоо', 'qty', 'ширхэг', 'хэмжээ'],
}

def _rows_from_upload(content: bytes, ext: str):
    if ext in ('xlsx', 'xlsm'):
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            if r is None or all(c is None or str(c).strip() == '' for c in r):
                continue
            rows.append(list(r))
        wb.close()
        return rows
    return _rows_from_file(content, ext)

def _sheets_from_upload(content: bytes, ext: str):
    """Return a list of row-lists — one per worksheet for Excel, or one for CSV/JSON."""
    if ext in ('xlsx', 'xlsm'):
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for r in ws.iter_rows(values_only=True):
                if r is None or all(c is None or str(c).strip() == '' for c in r):
                    continue
                rows.append(list(r))
            if rows:
                sheets.append(rows)
        wb.close()
        return sheets if sheets else [[]]
    return [_rows_from_file(content, ext)]

def _map_headers(header_row):
    idx = {}
    for i, h in enumerate(header_row):
        hl = str(h if h is not None else '').strip().lower()
        if not hl:
            continue
        for field, keys in HEADER_KEYS.items():
            if field not in idx and any(k in hl for k in keys):
                idx[field] = i
    return idx

def _cell(row, idx, field):
    i = idx.get(field)
    return row[i] if (i is not None and i < len(row)) else None

def _to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[^\d.\-]', '', str(v if v is not None else ''))
    if s in ('', '-', '.', '--'):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def _norm_date(v, default):
    if v is None or str(v).strip() == '':
        return default
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()[:10]
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%Y.%m.%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return default

def _norm_type(v, amount):
    s = str(v if v is not None else '').strip().lower()
    if 'орлог' in s or s.startswith('inc') or s in ('in', '+'):
        return 'income'
    if 'зарлаг' in s or s.startswith('exp') or s.startswith('out') or s == '-':
        return 'outcome'
    return 'income' if amount >= 0 else 'outcome'

def _guess_numeric_col(rows, exclude=None):
    ncols = max((len(r) for r in rows), default=0)
    best, best_count = None, 0
    for ci in range(ncols):
        if ci == exclude:
            continue
        cnt = 0
        for r in rows[1:60]:
            if ci < len(r):
                v = r[ci]
                if isinstance(v, (int, float)):
                    cnt += 1
                else:
                    s = str(v if v is not None else '').strip()
                    if s and re.fullmatch(r'[\d.,\-\s₮]+', s):
                        cnt += 1
        if cnt > best_count:
            best_count, best = cnt, ci
    return best if best_count >= 1 else None

def _find_header_row(rows):
    """Find the real header row, skipping title rows like 'Сангийн материал'."""
    allkeys = [k for keys in HEADER_KEYS.values() for k in keys]
    best_i, best_score = 0, 0
    for i, r in enumerate(rows[:12]):
        joined = ' '.join(str(c if c is not None else '').lower() for c in r)
        nonempty = sum(1 for c in r if str(c if c is not None else '').strip())
        score = sum(1 for k in allkeys if k in joined)
        if nonempty >= 2 and score > best_score:
            best_score, best_i = score, i
    return best_i if best_score >= 2 else 0

def _parse_import(rows, kind, default_date):
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Толгой мөр + дор хаяж 1 мөр хэрэгтэй / Need a header row + at least 1 data row")
    hi = _find_header_row(rows)
    header = rows[hi]
    data_rows = rows[hi + 1:]
    idx = _map_headers(header)
    records = []
    if kind == 'ledger':
        if 'amount' not in idx:
            guess = _guess_numeric_col(rows[hi:], exclude=idx.get('date'))
            if guess is None:
                raise HTTPException(status_code=400, detail="Мөнгөн дүнгийн багана олдсонгүй (Дүн / Үнэ / Amount) / Amount column not found")
            idx['amount'] = guess
        for r in data_rows:
            amt = _to_num(_cell(r, idx, 'amount'))
            if amt == 0:
                continue
            records.append({
                'date': _norm_date(_cell(r, idx, 'date'), default_date),
                'type': _norm_type(_cell(r, idx, 'type'), amt), 'amount': abs(amt),
                'category': str(_cell(r, idx, 'category') or '').strip(),
                'description': str(_cell(r, idx, 'description') or '').strip(),
            })
    elif kind == 'materials':
        if 'name' not in idx:
            idx['name'] = 0  # assume the first column holds the material name
        for r in data_rows:
            name = str(_cell(r, idx, 'name') or '').strip()
            if not name:
                continue
            price = _to_num(_cell(r, idx, 'price'))
            q = _cell(r, idx, 'quantity')
            qty = _to_num(q) if (q is not None and str(q).strip() != '') else None
            if price == 0 and qty is None:
                continue  # skip section/category label rows (e.g. "Цэвэр ус")
            records.append({
                'name': name, 'category': str(_cell(r, idx, 'category') or '').strip(),
                'unit': (str(_cell(r, idx, 'unit') or '').strip() or 'ш'),
                'price': price, 'supplier': str(_cell(r, idx, 'supplier') or '').strip(),
                'quantity': qty,
            })
    else:
        raise HTTPException(status_code=400, detail="kind must be 'ledger' or 'materials'")
    return records, list(idx.keys())

@api_router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    kind: str = Form(...),
    month: str = Form(""),
    confirm: str = Form("false"),
    authorization: Optional[str] = Header(None),
):
    user = await get_current_user(authorization)
    _rate_limit("imp:" + user["user_id"], 60, 60)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Хоосон файл / Empty file")
    if len(content) > 8_000_000:
        raise HTTPException(status_code=413, detail="Файл хэт том / File too large")
    fname = file.filename or "file"
    ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
    default_date = (month + "-15") if re.match(r'^\d{4}-\d{2}$', month or '') else datetime.now().strftime("%Y-%m-%d")
    sheets = _sheets_from_upload(content, ext)
    records, columns, errors = [], [], []
    for srows in sheets:
        if not srows or len(srows) < 2:
            continue
        try:
            recs, cols = _parse_import(srows, kind, default_date)
            records.extend(recs)
            if cols:
                columns = cols
        except HTTPException as e:
            errors.append(e.detail)
    if not records:
        raise HTTPException(status_code=400, detail=(errors[0] if errors else "Импортлох мөр олдсонгүй / No rows to import"))

    if str(confirm).lower() != "true":
        return {"preview": records[:12], "total": len(records), "columns": columns, "kind": kind}

    now = datetime.now(timezone.utc)
    if kind == 'ledger':
        docs = [{
            "transaction_id": f"txn_{uuid.uuid4().hex[:12]}", "user_id": user["user_id"],
            "type": rec['type'], "category": rec['category'], "amount": rec['amount'],
            "description": rec['description'], "date": rec['date'], "created_at": now,
            "material_id": "", "material_name": "", "quantity": None, "unit_price": None,
        } for rec in records]
        if docs:
            await db.transactions.insert_many(docs)
        return {"imported": len(docs), "kind": kind}
    else:
        mat_count, stock_count = 0, 0
        for rec in records:
            existing = await db.materials.find_one({"user_id": user["user_id"], "name": rec['name']}, {"_id": 0})
            if existing:
                mid = existing["material_id"]
                upd = {}
                if rec['price']:
                    upd['price'] = rec['price']
                if rec['supplier']:
                    upd['supplier'] = rec['supplier']
                if rec['category']:
                    upd['category'] = rec['category']
                if upd:
                    await db.materials.update_one({"material_id": mid}, {"$set": upd})
            else:
                mid = f"mat_{uuid.uuid4().hex[:12]}"
                await db.materials.insert_one({
                    "material_id": mid, "user_id": user["user_id"], "name": rec['name'],
                    "category": rec['category'], "supplier": rec['supplier'], "price": rec['price'],
                    "unit": rec['unit'], "min_stock": 0, "description": "", "image_url": "", "created_at": now,
                })
                mat_count += 1
            if rec['quantity']:
                d = datetime.strptime(default_date, "%Y-%m-%d")
                await db.stock_entries.insert_one({
                    "entry_id": f"stock_{uuid.uuid4().hex[:12]}", "user_id": user["user_id"],
                    "material_id": mid, "material_name": rec['name'], "unit": rec['unit'],
                    "quantity": rec['quantity'], "date": default_date, "month": d.month, "year": d.year,
                    "created_at": now,
                })
                stock_count += 1
        return {"imported": len(records), "materials_added": mat_count, "stock_added": stock_count, "kind": kind}

@api_router.get("/import/template")
async def import_template(kind: str, authorization: Optional[str] = Header(None)):
    await get_current_user(authorization)
    wb = Workbook(); ws = wb.active
    fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    if kind == 'ledger':
        ws.title = "Orlogo-Zarlaga"
        headers = ["Огноо/Date", "Төрөл/Type", "Дүн/Amount", "Ангилал/Category", "Тайлбар/Description"]
        sample = [["2026-06-05", "Орлого/income", 5000000, "Борлуулалт", "Гэрээ"],
                  ["2026-06-06", "Зарлага/expense", 555000, "Материал", "Цемент"]]
    else:
        kind = 'materials'
        ws.title = "Material"
        headers = ["Нэр/Name", "Ангилал/Category", "Нэгж/Unit", "Нэгж үнэ/Price", "Нийлүүлэгч/Supplier", "Тоо хэмжээ/Quantity"]
        sample = [["Цемент 500", "Цемент", "уут", 18500, "Хөтөл", 40],
                  ["Арматур 12мм", "Төмөр", "м", 3200, "Дархан", 455]]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(sample, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = w + 4
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="template_{kind}.xlsx"'})

@api_router.get("/report/monthly")
async def monthly_report(month: int, year: int, authorization: Optional[str] = Header(None)):
    user = await get_current_user(authorization)
    uid_q = {"user_id": user["user_id"]}
    start = f"{year}-{month:02d}-01"
    end = f"{year+1}-01-01" if month == 12 else f"{year}-{(month+1):02d}-01"
    txns = await db.transactions.find({**uid_q, "date": {"$gte": start, "$lt": end}}, {"_id": 0}).sort("date", 1).to_list(20000)
    stock = await db.stock_entries.find({**uid_q, "month": month, "year": year}, {"_id": 0}).sort("date", 1).to_list(20000)
    materials = await db.materials.find(uid_q, {"_id": 0}).sort("name", 1).to_list(5000)
    loans = await db.loans.find(uid_q, {"_id": 0}).to_list(2000)

    wb = Workbook()
    fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    hf = Font(bold=True, color="FFFFFF")

    def sheet(title, headers, data):
        ws = wb.create_sheet(title)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h); cell.fill = fill; cell.font = hf; cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(data, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, 50)
        ws.freeze_panes = "A2"

    inc = sum(t['amount'] for t in txns if t['type'] == 'income')
    exp = sum(t['amount'] for t in txns if t['type'] == 'outcome')
    # Break income & expense down by category (Цалин, Тээвэр, Материал, Түрээс ...)
    exp_by, inc_by = {}, {}
    for t in txns:
        cat = (str(t.get('category') or '').strip() or '— (ангилалгүй)')
        if t['type'] == 'income':
            inc_by[cat] = inc_by.get(cat, 0) + t['amount']
        else:
            exp_by[cat] = exp_by.get(cat, 0) + t['amount']
    summ = wb.active; summ.title = "Хураангуй-Summary"
    co = user.get('company') or ''
    srows = [
        [co], [f"{year}-{month:02d} сарын тайлан / Monthly report"], [],
        ["Нийт орлого / Total income", inc], ["Нийт зарлага / Total expense", exp], ["Зөрүү / Net", inc - exp], [],
        ["── ЗАРЛАГА ангиллаар / EXPENSE by category ──"],
    ]
    for cat, v in sorted(exp_by.items(), key=lambda x: -x[1]):
        srows.append([cat, v])
    srows += [[], ["── ОРЛОГО ангиллаар / INCOME by category ──"]]
    for cat, v in sorted(inc_by.items(), key=lambda x: -x[1]):
        srows.append([cat, v])
    srows += [[], ["Гүйлгээ / Transactions", len(txns)], ["Тооллого / Stock entries", len(stock)],
              ["Материал / Materials", len(materials)], ["Зээл / Loans", len(loans)]]
    for ri, row in enumerate(srows, 1):
        for ci, val in enumerate(row, 1):
            summ.cell(row=ri, column=ci, value=val)
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        if a not in (None, '') and (b is None or b == ''):
            summ.cell(row=ri, column=1).font = Font(bold=True, size=13 if ri == 1 else 11)
    summ.column_dimensions['A'].width = 36; summ.column_dimensions['B'].width = 18

    sheet("Орлого-Зарлага", ["Огноо", "Төрөл", "Ангилал", "Утга", "Дүн (₮)"],
          [[t['date'], ("Орлого" if t['type'] == 'income' else "Зарлага"), t.get('category', ''), t.get('description', ''), t['amount']] for t in txns])
    sheet("Тооллого-Stock", ["Огноо", "Материал", "Нэгж", "Тоо хэмжээ"],
          [[s['date'], s['material_name'], s.get('unit', ''), s['quantity']] for s in stock])
    sheet("Материал-Materials", ["Нэр", "Ангилал", "Нэгж", "Нэгж үнэ (₮)", "Нийлүүлэгч"],
          [[m['name'], m.get('category', ''), m.get('unit', ''), m.get('price', 0), m.get('supplier', '')] for m in materials])
    sheet("Зээл-Loans", ["Зээлдүүлэгч", "Үлдэгдэл (₮)", "Хүү %", "Төлөлт", "Дараагийн төлбөр", "Төлөв"],
          [[l['name'], l['amount'], l.get('interest_rate', 0), l.get('payment_amount', 0), l.get('next_payment_date', ''), l.get('status', '')] for l in loans])

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    safe = "".join(c if (ord(c) < 128 and c not in '\\/:*?"<>|') else "_" for c in (co or "report")).strip() or "report"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{safe}_{year}-{month:02d}.xlsx"'})

# ==================== APP WIRING ====================

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in CORS_ORIGINS.split(",")] if CORS_ORIGINS != "*" else ["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the frontend from the same service (single-deploy hosting). The /api routes above
# are matched first; everything else is served as static files (index.html for "/").
FRONTEND_DIR = (ROOT_DIR.parent / "frontend")
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")

@app.on_event("startup")
async def startup_indexes():
    try:
        await db.users.create_index("email", unique=True)
        await db.users.create_index("user_id", unique=True)
        await db.materials.create_index("material_id", unique=True)
        await db.materials.create_index("user_id")
        await db.stock_entries.create_index("entry_id", unique=True)
        await db.stock_entries.create_index([("user_id", 1), ("year", 1), ("month", 1)])
        await db.transactions.create_index("transaction_id", unique=True)
        await db.transactions.create_index([("user_id", 1), ("date", 1)])
        await db.loans.create_index("loan_id", unique=True)
        await db.loans.create_index([("user_id", 1), ("next_payment_date", 1)])
        logger.info("Indexes ready")
    except Exception as e:
        logger.error(f"Index error: {e}")

@app.on_event("shutdown")
async def shutdown():
    client.close()
